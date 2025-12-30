"""
MST (Microscale Thermophoresis) data parsers for NanoTemper instruments.

Supports parsing of NanoTemper XLSX and CSV files with dose-response data
and automatic calculation of normalized fluorescence (Fnorm) values.
"""

from __future__ import annotations

import csv
import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import numpy as np

logger = logging.getLogger(__name__)


@dataclass
class DosePoint:
    """Single dose-response data point from MST experiment."""
    
    concentration: float  # Ligand concentration (nM, μM, etc.)
    fnorm: float  # Normalized fluorescence (‰)
    fnorm_error: float  # Standard error of Fnorm
    cold_mean: float  # Mean cold fluorescence
    hot_mean: float  # Mean hot fluorescence
    capillary: int = 1  # Capillary number
    outlier: bool = False  # Outlier flag
    
    
@dataclass
class MSTData:
    """Complete MST dataset from instrument file."""
    
    instrument: str  # Instrument model (e.g., "NanoTemper Monolith NT.115")
    capillary_type: str  # Capillary type (e.g., "Premium", "Standard")
    excitation_power: float  # LED excitation power (%)
    mst_power: float  # MST laser power (%)
    dose_points: List[DosePoint]  # Dose-response data points
    metadata: Dict[str, Any]  # Additional metadata from file
    target_name: Optional[str] = None  # Target protein name
    ligand_name: Optional[str] = None  # Ligand name
    buffer: Optional[str] = None  # Buffer composition
    temperature: float = 25.0  # Experiment temperature (°C)


def calculate_fnorm(cold: np.ndarray, hot: np.ndarray) -> float:
    """
    Calculate normalized fluorescence (Fnorm) from cold and hot fluorescence values.
    
    Fnorm = (F_hot - F_cold) / F_cold * 1000 (in ‰)
    
    Args:
        cold: Cold fluorescence values
        hot: Hot fluorescence values
        
    Returns:
        Normalized fluorescence in per mille (‰)
        
    Raises:
        ValueError: If arrays are empty or cold fluorescence is zero
    """
    if len(cold) == 0 or len(hot) == 0:
        raise ValueError("Empty fluorescence arrays")
    
    if len(cold) != len(hot):
        raise ValueError("Cold and hot arrays must have same length")
    
    cold_mean = np.mean(cold)
    hot_mean = np.mean(hot)
    
    if cold_mean == 0:
        raise ValueError("Cold fluorescence cannot be zero")
    
    fnorm = (hot_mean - cold_mean) / cold_mean * 1000.0
    
    logger.debug(f"Calculated Fnorm: {fnorm:.2f}‰ (cold={cold_mean:.0f}, hot={hot_mean:.0f})")
    
    return fnorm


def parse_nanotemper_xlsx(path: Union[str, Path]) -> MSTData:
    """
    Parse NanoTemper XLSX export file.
    
    Expected format:
    - Excel file with multiple sheets
    - "Analysis" sheet contains dose-response data
    - "Raw Data" sheet contains individual measurements
    - Metadata in first rows of Analysis sheet
    
    Args:
        path: Path to NanoTemper XLSX file
        
    Returns:
        MSTData object with parsed dose-response data
        
    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file format is invalid
        ImportError: If openpyxl is not available
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"MST file not found: {path}")
    
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required for XLSX parsing. Install with: pip install openpyxl")
    
    logger.info(f"Parsing NanoTemper XLSX: {path}")
    
    try:
        workbook = openpyxl.load_workbook(path, data_only=True)
        
        # Look for Analysis sheet first, then any sheet with dose-response data
        analysis_sheet = None
        if "Analysis" in workbook.sheetnames:
            analysis_sheet = workbook["Analysis"]
        elif "Dose Response" in workbook.sheetnames:
            analysis_sheet = workbook["Dose Response"]
        else:
            # Use first sheet as fallback
            analysis_sheet = workbook.active
        
        if analysis_sheet is None:
            raise ValueError("No suitable sheet found in XLSX file")
        
        # Parse metadata from first rows
        instrument = "NanoTemper Monolith"
        capillary_type = "Standard"
        excitation_power = 20.0
        mst_power = 40.0
        target_name = None
        ligand_name = None
        buffer = None
        temperature = 25.0
        metadata = {}
        
        # Scan first 20 rows for metadata
        for row in range(1, 21):
            cell_a = analysis_sheet.cell(row, 1).value
            cell_b = analysis_sheet.cell(row, 2).value
            
            if cell_a is None:
                continue
                
            cell_a_str = str(cell_a).strip().lower()
            cell_b_str = str(cell_b).strip() if cell_b is not None else ""
            
            if 'instrument' in cell_a_str:
                instrument = cell_b_str or instrument
            elif 'capillary' in cell_a_str:
                capillary_type = cell_b_str or capillary_type
            elif 'excitation' in cell_a_str or 'led' in cell_a_str:
                try:
                    excitation_power = float(re.findall(r'[\d\.]+', cell_b_str)[0])
                except (IndexError, ValueError):
                    pass
            elif 'mst' in cell_a_str or 'laser' in cell_a_str:
                try:
                    mst_power = float(re.findall(r'[\d\.]+', cell_b_str)[0])
                except (IndexError, ValueError):
                    pass
            elif 'target' in cell_a_str or 'protein' in cell_a_str:
                target_name = cell_b_str
            elif 'ligand' in cell_a_str or 'compound' in cell_a_str:
                ligand_name = cell_b_str
            elif 'buffer' in cell_a_str:
                buffer = cell_b_str
            elif 'temperature' in cell_a_str:
                try:
                    temperature = float(re.findall(r'[\d\.]+', cell_b_str)[0])
                except (IndexError, ValueError):
                    pass
        
        # Find data table (look for concentration column)
        data_start_row = None
        conc_col = None
        fnorm_col = None
        error_col = None
        cold_col = None
        hot_col = None
        
        for row in range(1, analysis_sheet.max_row + 1):
            for col in range(1, min(analysis_sheet.max_column + 1, 20)):
                cell_value = analysis_sheet.cell(row, col).value
                if cell_value is None:
                    continue
                    
                cell_str = str(cell_value).strip().lower()
                
                if 'concentration' in cell_str or 'conc' in cell_str:
                    data_start_row = row
                    conc_col = col
                elif 'fnorm' in cell_str or 'normalized' in cell_str:
                    fnorm_col = col
                elif 'error' in cell_str or 'std' in cell_str:
                    error_col = col
                elif 'cold' in cell_str:
                    cold_col = col
                elif 'hot' in cell_str:
                    hot_col = col
            
            if data_start_row is not None:
                break
        
        if data_start_row is None or conc_col is None:
            raise ValueError("Concentration column not found in XLSX file")
        
        # Parse dose-response data
        dose_points = []
        
        for row in range(data_start_row + 1, analysis_sheet.max_row + 1):
            conc_cell = analysis_sheet.cell(row, conc_col).value
            if conc_cell is None:
                continue
            
            try:
                concentration = float(conc_cell)
                
                # Get Fnorm value
                fnorm = 0.0
                if fnorm_col is not None:
                    fnorm_cell = analysis_sheet.cell(row, fnorm_col).value
                    if fnorm_cell is not None:
                        fnorm = float(fnorm_cell)
                
                # Get error
                fnorm_error = 0.0
                if error_col is not None:
                    error_cell = analysis_sheet.cell(row, error_col).value
                    if error_cell is not None:
                        fnorm_error = float(error_cell)
                
                # Get cold/hot fluorescence
                cold_mean = 0.0
                hot_mean = 0.0
                
                if cold_col is not None:
                    cold_cell = analysis_sheet.cell(row, cold_col).value
                    if cold_cell is not None:
                        cold_mean = float(cold_cell)
                
                if hot_col is not None:
                    hot_cell = analysis_sheet.cell(row, hot_col).value
                    if hot_cell is not None:
                        hot_mean = float(hot_cell)
                
                # Calculate Fnorm if not provided but cold/hot are available
                if fnorm == 0.0 and cold_mean > 0 and hot_mean > 0:
                    fnorm = calculate_fnorm(np.array([cold_mean]), np.array([hot_mean]))
                
                dose_point = DosePoint(
                    concentration=concentration,
                    fnorm=fnorm,
                    fnorm_error=fnorm_error,
                    cold_mean=cold_mean,
                    hot_mean=hot_mean,
                    capillary=1,
                    outlier=False
                )
                dose_points.append(dose_point)
                
            except (ValueError, TypeError) as e:
                logger.warning(f"Skipping invalid data row {row}: {e}")
                continue
        
        if not dose_points:
            raise ValueError("No valid dose-response data found in XLSX file")
        
        # Sort by concentration
        dose_points.sort(key=lambda x: x.concentration)
        
        mst_data = MSTData(
            instrument=instrument,
            capillary_type=capillary_type,
            excitation_power=excitation_power,
            mst_power=mst_power,
            dose_points=dose_points,
            metadata=metadata,
            target_name=target_name,
            ligand_name=ligand_name,
            buffer=buffer,
            temperature=temperature
        )
        
        logger.info(f"Successfully parsed {len(dose_points)} dose points from {path}")
        return mst_data
        
    except Exception as e:
        logger.error(f"Failed to parse NanoTemper XLSX {path}: {e}")
        raise ValueError(f"Invalid NanoTemper XLSX format: {e}") from e


def parse_mst_csv(path: Union[str, Path]) -> MSTData:
    """
    Parse MST CSV export file.
    
    Expected format:
    - Comma-separated values
    - Header rows with metadata
    - Data columns: Concentration, Fnorm, Error, Cold, Hot
    
    Args:
        path: Path to MST CSV file
        
    Returns:
        MSTData object with parsed dose-response data
        
    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file format is invalid
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"MST file not found: {path}")
    
    logger.info(f"Parsing MST CSV: {path}")
    
    try:
        with open(path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # Parse metadata from header
        instrument = "NanoTemper"
        capillary_type = "Standard"
        excitation_power = 20.0
        mst_power = 40.0
        target_name = None
        ligand_name = None
        buffer = None
        temperature = 25.0
        metadata = {}
        
        data_start = 0
        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue
            
            # Check if this looks like a header line (contains column names)
            if not data_start and ('concentration' in line.lower() or 'fnorm' in line.lower()):
                data_start = i
                break
            # Check if this looks like a data line (starts with number and has commas)
            elif re.match(r'^[\d\-\.]', line) and ',' in line:
                if data_start == 0:
                    # No header found, assume first data line and create default header
                    parts = line.split(',')
                    if len(parts) >= 2:
                        lines.insert(i, "Concentration,Fnorm,Error,Cold,Hot")  # Default header
                        data_start = i
                        break
                else:
                    data_start = i
                    break
            
            # Parse metadata
            if ':' in line:
                key, value = line.split(':', 1)
                key = key.strip().lower()
                value = value.strip()
                metadata[key] = value
                
                if 'instrument' in key:
                    instrument = value
                elif 'capillary' in key:
                    capillary_type = value
                elif 'excitation' in key or 'led' in key:
                    try:
                        excitation_power = float(re.findall(r'[\d\.]+', value)[0])
                    except (IndexError, ValueError):
                        pass
                elif 'mst' in key or 'laser' in key:
                    try:
                        mst_power = float(re.findall(r'[\d\.]+', value)[0])
                    except (IndexError, ValueError):
                        pass
                elif 'target' in key or 'protein' in key:
                    target_name = value
                elif 'ligand' in key or 'compound' in key:
                    ligand_name = value
                elif 'buffer' in key:
                    buffer = value
                elif 'temperature' in key:
                    try:
                        temperature = float(re.findall(r'[\d\.]+', value)[0])
                    except (IndexError, ValueError):
                        pass
        
        # Parse CSV data
        csv_reader = csv.reader(lines[data_start:])
        rows = list(csv_reader)
        
        if len(rows) < 2:
            raise ValueError("Insufficient data in CSV file")
        
        # Determine column indices
        header = [col.strip().lower() for col in rows[0]]
        conc_col = None
        fnorm_col = None
        error_col = None
        cold_col = None
        hot_col = None
        
        for i, col in enumerate(header):
            if 'concentration' in col or 'conc' in col or col == 'concentration':
                conc_col = i
            elif 'fnorm' in col or 'normalized' in col:
                fnorm_col = i
            elif 'error' in col or 'std' in col:
                error_col = i
            elif 'cold' in col:
                cold_col = i
            elif 'hot' in col:
                hot_col = i
        
        if conc_col is None:
            raise ValueError("Concentration column not found in CSV")
        
        # Parse data rows
        dose_points = []
        
        for row in rows[1:]:
            if len(row) <= conc_col:
                continue
            
            try:
                concentration = float(row[conc_col])
                
                # Get Fnorm value
                fnorm = 0.0
                if fnorm_col is not None and len(row) > fnorm_col:
                    fnorm = float(row[fnorm_col])
                
                # Get error
                fnorm_error = 0.0
                if error_col is not None and len(row) > error_col:
                    try:
                        fnorm_error = float(row[error_col])
                    except ValueError:
                        pass
                
                # Get cold/hot fluorescence
                cold_mean = 0.0
                hot_mean = 0.0
                
                if cold_col is not None and len(row) > cold_col:
                    try:
                        cold_mean = float(row[cold_col])
                    except ValueError:
                        pass
                
                if hot_col is not None and len(row) > hot_col:
                    try:
                        hot_mean = float(row[hot_col])
                    except ValueError:
                        pass
                
                # Calculate Fnorm if not provided but cold/hot are available
                if fnorm == 0.0 and cold_mean > 0 and hot_mean > 0:
                    fnorm = calculate_fnorm(np.array([cold_mean]), np.array([hot_mean]))
                
                dose_point = DosePoint(
                    concentration=concentration,
                    fnorm=fnorm,
                    fnorm_error=fnorm_error,
                    cold_mean=cold_mean,
                    hot_mean=hot_mean,
                    capillary=1,
                    outlier=False
                )
                dose_points.append(dose_point)
                
            except (ValueError, IndexError) as e:
                logger.warning(f"Skipping invalid data row: {row[:3]}... ({e})")
                continue
        
        if not dose_points:
            raise ValueError("No valid dose-response data found in CSV file")
        
        # Sort by concentration
        dose_points.sort(key=lambda x: x.concentration)
        
        mst_data = MSTData(
            instrument=instrument,
            capillary_type=capillary_type,
            excitation_power=excitation_power,
            mst_power=mst_power,
            dose_points=dose_points,
            metadata=metadata,
            target_name=target_name,
            ligand_name=ligand_name,
            buffer=buffer,
            temperature=temperature
        )
        
        logger.info(f"Successfully parsed {len(dose_points)} dose points from {path}")
        return mst_data
        
    except Exception as e:
        logger.error(f"Failed to parse MST CSV {path}: {e}")
        raise ValueError(f"Invalid MST CSV format: {e}") from e


def extract_dose_response(data: MSTData) -> List[DosePoint]:
    """
    Extract dose-response data points from MSTData object.
    
    Args:
        data: MSTData object
        
    Returns:
        List of DosePoint objects sorted by concentration
    """
    return sorted(data.dose_points, key=lambda x: x.concentration)


__all__ = [
    "DosePoint",
    "MSTData",
    "parse_nanotemper_xlsx",
    "parse_mst_csv",
    "extract_dose_response",
    "calculate_fnorm",
]
